import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook

folder1 = "output_by_subject"
 
if not os.path.exists(folder1): 
        os.makedirs(folder1)
folder2 = "output_individual_roll"
 
if not os.path.exists(folder2): 
        os.makedirs(folder2)

def output_by_subject():
    head=['rollno','register_sem','subno','sub_type']
    with open(r'tut04\regtable_old.csv','r') as info:
        cinfo=csv.reader(info)
        for words in cinfo:
            given_data=[]
            given_data.append(words[0])
            given_data.append(words[1])
            given_data.append(words[3])
            given_data.append(words[-1])
            if (given_data[2] =="subno"):continue
            subno='{}.xlsx'.format(given_data[2])
            path='./output_by_subject/'+subno
        
            if(os.path.isfile(path)): 
                 book=load_workbook(r'output_by_subject\\{}.xlsx'.format(given_data[2]))
                 sheet1=book.active
                 sheet1.append(given_data)
                 book.save(r'output_by_subject\\{}.xlsx'.format(given_data[2]))
                 
            else: 
                    book=Workbook()
                    sheet1=book.active
                    sheet1.append(head)
                    sheet1.append(given_data)
                    book.save(f'output_by_subject\\{given_data[2]}.xlsx')
    return
 
def output_individual_roll():
    head=['rollno','register_sem','subno','sub_type']
    with open(r'tut04\regtable_old.csv','r') as info:
        cinfo=csv.reader(info)
        for words in cinfo:
            given_data=[]
            given_data.append(words[0])
            given_data.append(words[1])
            given_data.append(words[3])
            given_data.append(words[-1])
            if (given_data[0] =="rollno"):continue
            rollno='{}.xlsx'.format(given_data[0])
            path='./output_individual_roll/'+rollno
            if(os.path.isfile(path)): 
                 book=load_workbook(r'output_individual_roll\\{}.xlsx'.format(given_data[0]))
                 sheet2=book.active
                 sheet2.append(given_data)
                 book.save(r'output_individual_roll\\{}.xlsx'.format(given_data[0]))
                 
                    
            else: 
                    book=Workbook()
                    sheet2=book.active
                    sheet2.append(head)
                    sheet2.append(given_data)
                    book.save(f'output_individual_roll\\{given_data[0]}.xlsx')
    return
 
output_by_subject()
output_individual_roll()