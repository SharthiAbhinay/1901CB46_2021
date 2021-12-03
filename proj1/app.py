import pandas as pd
import os
import openpyxl 
import numpy as np
from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,NamedStyle,Alignment,Side,Border
from flask import Flask,render_template,request,url_for,redirect,send_file,send_from_directory
from werkzeug.utils import secure_filename
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
app = Flask(__name__)
app.config['SECRET_KEY'] = 'mysecret'
app.config['UPLOAD_FOLDER']="C:\\Users\\shart\\OneDrive\\Desktop\\prctice\\static"
@app.route("/",methods=["POST","GET"])
def hello_world():
    if request.method=="POST":
        f=request.files["master_roll"]
        f.save(os.path.join(app.config['UPLOAD_FOLDER'],secure_filename(f.filename)))
        r=request.files["response"]
        r.save(os.path.join(app.config['UPLOAD_FOLDER'],secure_filename(r.filename)))
        c_marks=request.form.get("positivemarks",type=float)
        n_marks=request.form.get("negativemarks",type=float)
        if request.form['submit_button']== "Generate Roll No Wise Mark Sheet" :
            return redirect(url_for("genarate",cw=c_marks,wa=n_marks))
        elif  request.form['submit_button']== "Generate Concise Mark Sheet" :
            return redirect(url_for("concise",cw=c_marks,wa=n_marks))
        elif request.form['submit_button']== "Send Email" :  
            return redirect(url_for("email",cw=c_marks,wa=n_marks))     
    else:
        return render_template('GU.html')
@app.route("/genarate/<cw>,<wa>")       
def genarate(cw,wa):
    try:
        os.mkdir(r"C:\Users\shart\OneDrive\Desktop\prctice\static\Marksheet")
    except: 
        pass
        os.chdir(r"C:\Users\shart\OneDrive\Desktop\prctice\static\Marksheet")
    regi=pd.read_csv(r'C:\Users\shart\OneDrive\Desktop\prctice\static\master_roll.csv')
    resp=pd.read_csv(r'C:\Users\shart\OneDrive\Desktop\prctice\static\responses.csv')
    rollno= regi['roll'].values.tolist()
    ans=resp[resp['Roll Number']=="ANSWER"]
    correct_ans=ans.drop(labels=['Timestamp', 'Roll Number','Score','Name','IITP webmail','Phone (10 digit only)','Email address'],axis=1)
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True, size=12,name='Century',color='00000000')
    bd = Side(style='thin', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    highlight.alignment=Alignment(horizontal='center')
    g=NamedStyle(name="g")
    g.font=Font(name='Century',color='0000FF00',size=12)
    bd = Side(style='thin', color="000000")
    g.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    g.alignment=Alignment(horizontal='center')
    r=NamedStyle(name="r")
    r.font=Font(name='Century',color='00FF0000',size=12)
    bd = Side(style='thin', color="000000")
    r.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    r.alignment=Alignment(horizontal='center')
    m=NamedStyle(name="m")
    m.font=Font(name='Century',color='000000FF',size=12)
    bd = Side(style='thin', color="000000")
    m.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    m.alignment=Alignment(horizontal='center')
    n=NamedStyle(name="n")
    n.font=Font(name='Century',color='00000000',size=12)
    n.alignment=Alignment(horizontal='right')
    b=NamedStyle(name="b")
    b.font=Font(name='Century',color='00000000',size=12,bold=True)
    l=NamedStyle(name="l")
    l.font=Font(name='Century',color='00000000',size=12)
    bd = Side(style='thin', color="000000")
    l.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    l.alignment=Alignment(horizontal='center')
    if( (regi['roll']=="ANSWER").any()):
        for i in rollno:
            if i=="ANSWER":
                continue
            xlsx=openpyxl.Workbook()
            xlsx.add_named_style(highlight)
            xlsx.add_named_style(n)
            xlsx.add_named_style(m)
            xlsx.add_named_style(b)
            xlsx.add_named_style(r)
            xlsx.add_named_style(g)
            xlsx.add_named_style(l)
            sheet=xlsx.active
            sheet.column_dimensions['A'].width = 18.56
            sheet.column_dimensions['B'].width = 18.56
            sheet.column_dimensions['C'].width = 18.56
            sheet.column_dimensions['D'].width = 18.56
            sheet.column_dimensions['E'].width = 18.56
            for j in range(6,41):
                sheet.row_dimensions[j].height = 17
            
            
            name=i
            if ((resp['Roll Number']==i).any()):
                df=resp[resp['Roll Number']==i]
                img=Image(r'C:\Users\shart\OneDrive\Desktop\prctice\static\pic.png')
                img.width = 631.56
                img.height = 79.39
                sheet.add_image(img,"A1")
                sheet['C5']="Mark sheet"
                cell=sheet['C5']
                cell.font=Font(name='Century',size=18,color='00000000',bold=True,underline="single")
                sheet['A6']="Name"
                sheet['A6'].style='n'
                sheet['A7']="Roll Number"
                sheet['A7'].style='n'
                sheet['A10']="No."
                sheet['A10'].style='highlight'
                sheet['A11']="Marking"
                sheet['A11'].style='highlight'
                sheet['A12']="Total"
                sheet['A12'].style='highlight'
                sheet['A15']="Student Ans"
                sheet['A15'].style='highlight'
                sheet['B9']="Right"
                sheet['B9'].style='highlight'
                sheet['B15']="Correct Ans"
                sheet['B15'].style='highlight'
                sheet['C9']="Wrong"
                sheet['C9'].style='highlight'
                sheet['D6']="Exam"
                sheet['D6'].style='n'
                sheet['D9']="Not Attempt"
                sheet['D9'].style='highlight'
                sheet['D15']="Student Ans"
                sheet['D15'].style='highlight'
                sheet["E6"]="quiz"
                sheet['E6'].style='b'
                sheet['E9']="Max"
                sheet['E9'].style='highlight'
                sheet['E15']="Correct Ans"
                sheet['E15'].style='highlight'
                sheet['B6']=df.iloc[0]['Name']
                sheet['B6'].style='b'
                sheet['B7']=i
                sheet['B7'].style='b'
                fd=df.drop(labels=['Timestamp', 'Roll Number','Score','Name','IITP webmail','Phone (10 digit only)','Email address'],axis=1)
                correct=0
                notattempted=0
                wrong=0
                rows=16
                columnA=1
                columnB=2
                for key, value in fd.iteritems():
                    o=fd.iloc[0][key]
                    p=correct_ans.iloc[0][key]
                    if pd.isnull(fd.iloc[0][key]):
                        notattempted=notattempted+1
                    elif o==p:
                        correct=correct+1
                        sheet.cell(row=rows,column=columnA).value=o
                        sheet.cell(row=rows,column=columnA).style='g'
                    else:
                        wrong=wrong+1
                        sheet.cell(row=rows,column=columnA).value=o
                        sheet.cell(row=rows,column=columnA).style='r'
                    sheet.cell(row=rows,column=columnB).value=p
                    sheet.cell(row=rows,column=columnB).style='m'
                    rows=rows+1
                    if rows>40:
                        rows=16
                        columnA=4
                        columnB=5
                sheet['B10']=correct
                sheet['B10'].style='g'
                sheet['C10']=wrong
                sheet['c10'].style='r'
                sheet['D10']=notattempted
                sheet['D10'].style='l'
                sheet['E10']=correct+wrong+notattempted
                sheet['E10'].style='l'
                sheet['B11']=cw
                sheet['B11'].style='g'
                sheet['C11']=wa
                sheet['C11'].style='r'
                sheet['D11']=0
                sheet['D11'].style='l'
                sheet['B12']=correct*float(cw)
                sheet['B12'].style='g'
                sheet['C12']=wrong*float(wa)
                sheet['C12'].style='r'
                sheet['E11'].style='l'
                sheet['D12'].style='l'
                sheet['A9'].style='l'
                sheet['E12']=str((correct*float(cw))+(wrong*float(wa)))+"/"+str((wrong+correct+notattempted)*float(cw))
                sheet['E12'].style='m'
                    


            xlsx.save(f"C:\\Users\\shart\\OneDrive\\Desktop\\prctice\\static\\Marksheet\\{name}.xlsx")
        return redirect(url_for('logs'))
    else:
        return 'No roll number with ANSWER is present, Cannot Process!'
@app.route("/concise/<cw>,<wa>") 
def  concise(cw,wa):
    regi=pd.read_csv(r'C:\Users\shart\OneDrive\Desktop\prctice\static\master_roll.csv')
    resp=pd.read_csv(r'C:\Users\shart\OneDrive\Desktop\prctice\static\responses.csv')
    rollno= regi['roll'].values.tolist()
    ans=resp[resp['Roll Number']=="ANSWER"]
    correct_ans=ans.drop(labels=['Timestamp', 'Roll Number','Score','Name','IITP webmail','Phone (10 digit only)','Email address'],axis=1)
    l={}
    w={}
    b={}
    for i in rollno:
        d=resp[resp['Roll Number']==i]
        fd=d.drop(labels=['Timestamp', 'Roll Number','Score','Name','IITP webmail','Phone (10 digit only)','Email address'],axis=1)
        correct=0
        notattempted=0
        wrong=0
        for key, value in fd.iteritems():
            o=fd.iloc[0][key]
            p=correct_ans.iloc[0][key]
            if pd.isnull(fd.iloc[0][key]):
                notattempted=notattempted+1
            elif o==p:
                correct=correct+1        
            else:
                wrong=wrong+1
        l[i]=[correct,wrong,notattempted]
        h=str((correct*float(cw))+(wrong*float(wa)))+"/"+str((wrong+correct+notattempted)*float(cw))
        k=str(correct*float(cw))+"/"+str((wrong+correct+notattempted)*float(cw))
        w[i]=h
        b[i]=k
    df=pd.DataFrame()
    d=pd.DataFrame()
    o=pd.DataFrame()
    df["statusAns"]=""
    d["Score_After_Negative"]=""
    o["Google_Score"]=""
    i=0
    for key in l:
        df.at[i,'statusAns']=l[key]
        d.at[i,'Score_After_Negative']=w[key]
        o.at[i,'Google_Score']=b[key]
        i=i+1
    fd=pd.concat([resp,df],axis=1)
    newdf=pd.concat([fd,d],axis=1)
    newdf=pd.concat([newdf,o],axis=1)
    newdf=newdf.reindex(['Timestamp','Email address','Google_Score','IITP webmail','Name','Phone (10 digit only)', 'Score_After_Negative','Roll Number','Unnamed: 7','Unnamed: 8','Unnamed: 9','Unnamed: 10','Unnamed: 11','Unnamed: 12','Unnamed: 13','Unnamed: 14','Unnamed: 15','Unnamed: 16','Unnamed: 17','Unnamed: 18','Unnamed: 19','Unnamed: 20','Unnamed: 21','Unnamed: 22','Unnamed: 23','Unnamed: 24','Unnamed: 25','Unnamed: 26','Unnamed: 27','Unnamed: 28','Unnamed: 29','Unnamed: 30','Unnamed: 31','Unnamed: 32','Unnamed: 33','Unnamed: 34','statusAns'], axis=1)
    newdf.to_csv (r'C:\Users\shart\OneDrive\Desktop\prctice\static\ConciseMarksheet.csv', index = False, header=False)
    return redirect(url_for('file_downloads'))
@app.route("/email/<cw>,<wa>") 
def  email(cw,wa):
    try:
        os.mkdir(r"C:\Users\shart\OneDrive\Desktop\prctice\static\Excel_sheet")
    except: 
        pass
        os.chdir(r"C:\Users\shart\OneDrive\Desktop\prctice\static\Excel_sheet")
    regi=pd.read_csv(r'C:\Users\shart\OneDrive\Desktop\prctice\static\master_roll.csv')
    resp=pd.read_csv(r'C:\Users\shart\OneDrive\Desktop\prctice\static\responses.csv')
    rollno= regi['roll'].values.tolist()
    ans=resp[resp['Roll Number']=="ANSWER"]
    correct_ans=ans.drop(labels=['Timestamp', 'Roll Number','Score','Name','IITP webmail','Phone (10 digit only)','Email address'],axis=1)
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True, size=12,name='Century',color='00000000')
    bd = Side(style='thin', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    highlight.alignment=Alignment(horizontal='center')
    g=NamedStyle(name="g")
    g.font=Font(name='Century',color='0000FF00',size=12)
    bd = Side(style='thin', color="000000")
    g.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    g.alignment=Alignment(horizontal='center')
    r=NamedStyle(name="r")
    r.font=Font(name='Century',color='00FF0000',size=12)
    bd = Side(style='thin', color="000000")
    r.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    r.alignment=Alignment(horizontal='center')
    m=NamedStyle(name="m")
    m.font=Font(name='Century',color='000000FF',size=12)
    bd = Side(style='thin', color="000000")
    m.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    m.alignment=Alignment(horizontal='center')
    n=NamedStyle(name="n")
    n.font=Font(name='Century',color='00000000',size=12)
    n.alignment=Alignment(horizontal='right')
    b=NamedStyle(name="b")
    b.font=Font(name='Century',color='00000000',size=12,bold=True)
    l=NamedStyle(name="l")
    l.font=Font(name='Century',color='00000000',size=12)
    bd = Side(style='thin', color="000000")
    l.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    l.alignment=Alignment(horizontal='center')
    if( (regi['roll']=="ANSWER").any()):
        for i in rollno:
            if i=="ANSWER":
                continue
            xlsx=openpyxl.Workbook()
            xlsx.add_named_style(highlight)
            xlsx.add_named_style(n)
            xlsx.add_named_style(m)
            xlsx.add_named_style(b)
            xlsx.add_named_style(r)
            xlsx.add_named_style(g)
            xlsx.add_named_style(l)
            sheet=xlsx.active
            sheet.column_dimensions['A'].width = 18.56
            sheet.column_dimensions['B'].width = 18.56
            sheet.column_dimensions['C'].width = 18.56
            sheet.column_dimensions['D'].width = 18.56
            sheet.column_dimensions['E'].width = 18.56
            for j in range(6,41):
                sheet.row_dimensions[j].height = 17
            
            
            name=i
            if ((resp['Roll Number']==i).any()):
                df=resp[resp['Roll Number']==i]
                img=Image(r'C:\Users\shart\OneDrive\Desktop\prctice\static\pic.png')
                img.width = 631.56
                img.height = 79.39
                sheet.add_image(img,"A1")
                sheet['C5']="Mark sheet"
                cell=sheet['C5']
                cell.font=Font(name='Century',size=18,color='00000000',bold=True,underline="single")
                sheet['A6']="Name"
                sheet['A6'].style='n'
                sheet['A7']="Roll Number"
                sheet['A7'].style='n'
                sheet['A10']="No."
                sheet['A10'].style='highlight'
                sheet['A11']="Marking"
                sheet['A11'].style='highlight'
                sheet['A12']="Total"
                sheet['A12'].style='highlight'
                sheet['A15']="Student Ans"
                sheet['A15'].style='highlight'
                sheet['B9']="Right"
                sheet['B9'].style='highlight'
                sheet['B15']="Correct Ans"
                sheet['B15'].style='highlight'
                sheet['C9']="Wrong"
                sheet['C9'].style='highlight'
                sheet['D6']="Exam"
                sheet['D6'].style='n'
                sheet['D9']="Not Attempt"
                sheet['D9'].style='highlight'
                sheet['D15']="Student Ans"
                sheet['D15'].style='highlight'
                sheet["E6"]="quiz"
                sheet['E6'].style='b'
                sheet['E9']="Max"
                sheet['E9'].style='highlight'
                sheet['E15']="Correct Ans"
                sheet['E15'].style='highlight'
                sheet['B6']=df.iloc[0]['Name']
                sheet['B6'].style='b'
                sheet['B7']=i
                sheet['B7'].style='b'
                fd=df.drop(labels=['Timestamp', 'Roll Number','Score','Name','IITP webmail','Phone (10 digit only)','Email address'],axis=1)
                correct=0
                notattempted=0
                wrong=0
                rows=16
                columnA=1
                columnB=2
                for key, value in fd.iteritems():
                    o=fd.iloc[0][key]
                    p=correct_ans.iloc[0][key]
                    if pd.isnull(fd.iloc[0][key]):
                        notattempted=notattempted+1
                    elif o==p:
                        correct=correct+1
                        sheet.cell(row=rows,column=columnA).value=o
                        sheet.cell(row=rows,column=columnA).style='g'
                    else:
                        wrong=wrong+1
                        sheet.cell(row=rows,column=columnA).value=o
                        sheet.cell(row=rows,column=columnA).style='r'
                    sheet.cell(row=rows,column=columnB).value=p
                    sheet.cell(row=rows,column=columnB).style='m'
                    rows=rows+1
                    if rows>40:
                        rows=16
                        columnA=4
                        columnB=5
                sheet['B10']=correct
                sheet['B10'].style='g'
                sheet['C10']=wrong
                sheet['c10'].style='r'
                sheet['D10']=notattempted
                sheet['D10'].style='l'
                sheet['E10']=correct+wrong+notattempted
                sheet['E10'].style='l'
                sheet['B11']=cw
                sheet['B11'].style='g'
                sheet['C11']=wa
                sheet['C11'].style='r'
                sheet['D11']=0
                sheet['D11'].style='l'
                sheet['B12']=correct*float(cw)
                sheet['B12'].style='g'
                sheet['C12']=wrong*float(wa)
                sheet['C12'].style='r'
                sheet['E11'].style='l'
                sheet['D12'].style='l'
                sheet['A9'].style='l'
                sheet['E12']=str((correct*float(cw))+(wrong*float(wa)))+"/"+str((wrong+correct+notattempted)*float(cw))
                sheet['E12'].style='m'
            xlsx.save(f"C:\\Users\\shart\\OneDrive\\Desktop\\prctice\\static\\Excel_sheet\\{name}.xlsx")
        d=resp.applymap(str).groupby('Roll Number')['Email address'].apply(list).to_dict()
        df=resp.applymap(str).groupby('Roll Number')['IITP webmail'].apply(list).to_dict()
        for key in d:
            d[key].append(df[key])
        fromaddr = "sharthiabhinay@gmail.com"
        for key,value in d.items():
            for v in value:
                toaddr = v
                msg = MIMEMultipart()
                msg['From'] = fromaddr
                msg['To'] = ', '.join(toaddr)
                msg['Subject'] = "Quiz MarkSheet"
                body = "correct answer is marked:"+cw+" worng answer is marked:"+wa
                msg.attach(MIMEText(body, 'plain'))
                name=key
                filename = f"{name}.xlsx"
                attachment = open(f"C:\\Users\\shart\\OneDrive\\Desktop\\prctice\\static\\Excel_sheet\\{name}.xlsx", "rb")
                p = MIMEBase('application', 'octet-stream')
                p.set_payload((attachment).read())
                encoders.encode_base64(p)  
                p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
                msg.attach(p)
                s = smtplib.SMTP('smtp.gmail.com', 587)
                s.starttls()
                s.login(fromaddr, 'brkfmhfugkltfozi')
                text = msg.as_string()
                s.sendmail(fromaddr, toaddr, text)
        s.quit()
        return redirect(url_for('hello_world'))
    else:
        return 'No roll number with ANSWER is present, Cannot Process!'  
@app.route("/logs")
def logs():
    filenames = os.listdir(r'C:\Users\shart\OneDrive\Desktop\prctice\static\Marksheet')
    return render_template('logs.html', files=filenames)

@app.route('/logs/<path:filename>')
def log(filename):
    return send_from_directory(
        os.path.abspath(r'C:\Users\shart\OneDrive\Desktop\prctice\static\Marksheet'),
        filename,
        as_attachment=True
    )
@app.route('/file-downloads/')
def file_downloads():
	try:
		return render_template('downloads.html')
	except Exception as e:
		return str(e)
@app.route('/return_files/')
def return_files_tut():
	try:
	   return send_file(r'C:\Users\shart\OneDrive\Desktop\prctice\static\ConciseMarksheet.xlsx',attachment_filename='ConciseMarksheet.xlsx')
	except Exception as e:
		return str(e)
@app.route('/return_file_tut/<path:filename>')
def return_file_tut(filename):
	try:
	   return send_file(f'C:\\Users\\shart\\OneDrive\\Desktop\\prctice\\static\\Marksheet\\{filename}')
	except Exception as e:
		return str(e)  
@app.route('/files/<path:filename>',methods=["POST","GET"])
def files(filename):
    full_path = os.path.join(app.root_path, app.config['UPLOAD_FOLDER'])
    return send_from_directory(full_path, filename)

    
@app.route('/back')
def back():
    return redirect(url_for('hello_world'))   
if __name__ =="__main__":
     app.run(debug=True)